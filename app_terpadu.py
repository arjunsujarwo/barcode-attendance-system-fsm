"""
SISTEM INTEGRASI HRD & ABSENSI BARCODE (MANDIRI) - TIMING LOCKOUT VERSION
=====================================================================
Dibuat untuk: Keperluan Dokumentasi Proyek Akhir Sistem Informasi UBSI.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import cv2
from pyzbar.pyzbar import decode
import barcode
from barcode.writer import ImageWriter
import sys # Tambahkan import sys di bagian atas jika belum ada

def get_base_path():
    """Mendapatkan path absolut ke direktori tempat aplikasi berada, 
    baik saat dijalankan sebagai script maupun sebagai PyInstaller exe."""
    if getattr(sys, 'frozen', False):
        # Jika berjalan sebagai .exe, gunakan direktori file .exe berada
        return os.path.dirname(sys.executable)
    else:
        # Jika berjalan sebagai .py biasa
        return os.path.dirname(os.path.abspath(__file__))

class SistemHRDAndAbsensi:
    def __init__(self, root):
        """Inisialisasi konfigurasi dasar jendela aplikasi dan layout menu."""
        self.root = root
        self.root.title("HRD System & Attendance Barcode v2.1")
        self.root.geometry("700x600")
        self.root.configure(bg="#d4d0c8")
        
        # Variabel Kontrol Aplikasi
        self.after_id = None            
        
        # PERBAIKAN PATH UNTUK VERSI .EXE
        base_dir = get_base_path()
        self.nama_file_excel = os.path.join(base_dir, "database_absensi.xlsx")  
        self.folder_barcode = os.path.join(base_dir, "hasil_barcode")           
        self.is_camera_on = False       
        
        # Inisialisasi awal database lokal & folder barcode
        self.buat_atau_verifikasi_database()
        if not os.path.exists(self.folder_barcode):
            os.makedirs(self.folder_barcode)
        # === STYLING LAYOUT JADUL (Windows Classic Style) ===
        style = ttk.Style()
        style.theme_use('winnative')
        style.configure(".", background="#d4d0c8", foreground="black", font=("Arial", 10))
        style.configure("TNotebook", background="#d4d0c8", padding=2)
        style.configure("TNotebook.Tab", background="#d4d0c8", padding=[10, 4], font=("Arial", 9, "bold"))
        style.map("TNotebook.Tab", background=[("selected", "#b5b2ad")], foreground=[("selected", "black")])

        # === TOP BAR UTAMA ===
        self.top_bar = tk.Frame(root, bg="#0a246a", height=30)
        self.top_bar.pack(fill="x", padx=4, pady=(4, 0))
        self.lbl_title = tk.Label(self.top_bar, text="  Sistem Integrasi HRD & Absensi Barcode (Timing Lockout)", font=("MS Sans Serif", 10, "bold"), fg="white", bg="#0a246a")
        self.lbl_title.pack(side="left", pady=3)

        # === KONTROL MENU TAB (NOTEBOOK WIDGET) ===
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=6, pady=6)

        self.tab_absensi = tk.Frame(self.notebook, bg="#d4d0c8")
        self.tab_generator = tk.Frame(self.notebook, bg="#d4d0c8")
        self.tab_rekap = tk.Frame(self.notebook, bg="#d4d0c8")

        self.notebook.add(self.tab_absensi, text="  🖥️ MESIN ABSENSI  ")
        self.notebook.add(self.tab_generator, text="  ⚙️ MASTER KARYAWAN  ")
        self.notebook.add(self.tab_rekap, text="  📊 REKAP DATA BULANAN  ")

        self.buat_gui_tab_absensi()
        self.buat_gui_tab_generator()
        self.buat_gui_tab_rekap()

    # =========================================================================
    # LOGIKA MANAJEMEN DATABASE MULTI-SHEET
    # =========================================================================
    
    def buat_atau_verifikasi_database(self):
        """Membuat struktur database baru jika belum ada dengan sistem Multi-Sheet."""
        if not os.path.exists(self.nama_file_excel):
            wb = openpyxl.Workbook()
            
            # 1. Sheet Master Karyawan (Data Permanen)
            ws_karyawan = wb.active
            ws_karyawan.title = "Master_Karyawan"
            ws_karyawan.append(["ID_Karyawan", "Nama_Karyawan", "Tanggal_Terdaftar"])
            # Data awal/Dummy records
            ws_karyawan.append(["12345", "Arjun Sujarwo", datetime.now().strftime("%Y-%m-%d")])
            ws_karyawan.append(["67890", "Budi Santoso", datetime.now().strftime("%Y-%m-%d")])
            
            # 2. Sheet Log Absensi Harian (Data Terus Bertambah ke Bawah)
            ws_absensi = wb.create_sheet(title="Log_Absensi")
            ws_absensi.append(["Tanggal", "ID_Karyawan", "Nama", "Status_Terakhir", "Waktu_Masuk", "Waktu_Istirahat", "Waktu_Kembali", "Waktu_Pulang"])
            
            wb.save(self.nama_file_excel)

    def ambil_master_karyawan(self):
        """Membaca master data karyawan dari Excel ke RAM untuk proses validasi cepat."""
        karyawan_dict = {}
        if os.path.exists(self.nama_file_excel):
            wb = openpyxl.load_workbook(self.nama_file_excel, data_only=True)
            ws = wb["Master_Karyawan"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    karyawan_dict[str(row[0]).strip()] = row[1]
        return karyawan_dict

    def cari_atau_buat_log_hari_ini(self, id_karyawan, nama, tanggal_str):
        """Mencari data log absensi karyawan untuk hari ini serta mengembalikan semua data kolomnya."""
        wb = openpyxl.load_workbook(self.nama_file_excel)
        ws = wb["Log_Absensi"]
        
        target_row = None
        
        for r in range(2, ws.max_row + 1):
            tgl_cell = str(ws.cell(row=r, column=1).value).strip()
            id_cell = str(ws.cell(row=r, column=2).value).strip()
            
            if tgl_cell == tanggal_str and id_cell == id_karyawan:
                target_row = r
                break
        
        if target_row is None:
            ws.append([tanggal_str, id_karyawan, nama, "BELUM_MASUK", "", "", "", ""])
            target_row = ws.max_row
            wb.save(self.nama_file_excel)
            
        # Memuat ulang data spesifik baris log karyawan tersebut
        wb = openpyxl.load_workbook(self.nama_file_excel, data_only=True)
        ws = wb["Log_Absensi"]
        
        log_data = {
            "row_index": target_row,
            "status": ws.cell(row=target_row, column=4).value or "BELUM_MASUK",
            "waktu_masuk": ws.cell(row=target_row, column=5).value,
            "waktu_istirahat": ws.cell(row=target_row, column=6).value,
            "waktu_kembali": ws.cell(row=target_row, column=7).value,
            "waktu_pulang": ws.cell(row=target_row, column=8).value
        }
        return log_data

    def perbarui_log_absensi(self, row_index, kolom_ke, status_baru, waktu_scan):
        """Menuliskan update status dan waktu scan ke baris harian yang spesifik."""
        try:
            wb = openpyxl.load_workbook(self.nama_file_excel)
            ws = wb["Log_Absensi"]
            
            ws.cell(row=row_index, column=4, value=status_baru)
            ws.cell(row=row_index, column=kolom_ke, value=waktu_scan.strftime("%H:%M:%S"))
            
            wb.save(self.nama_file_excel)
            return True
        except PermissionError:
            messagebox.showerror("Permission Error", "Gagal menyimpan! Tutup file 'database_absensi.xlsx' terlebih dahulu.")
            return False

    # =========================================================================
    # MENU 1: HALAMAN TAB MESIN ABSENSI
    # =========================================================================
    
    def buat_gui_tab_absensi(self):
        frame_input = tk.Frame(self.tab_absensi, bg="#d4d0c8", bd=2, relief="groove")
        frame_input.pack(fill="x", padx=15, pady=15)
        
        tk.Label(frame_input, text="Nomor Karyawan :", font=("Arial", 11, "bold"), bg="#d4d0c8").pack(side="left", padx=(20, 5), pady=10)
        
        self.input_barcode = tk.Entry(frame_input, font=("Arial", 12), width=20, bd=2, relief="sunken")
        self.input_barcode.pack(side="left", padx=5, pady=10)
        self.input_barcode.focus()
        self.input_barcode.bind("<Return>", self.proses_absensi)

        tk.Button(frame_input, text="📷 Scan Kamera", font=("Arial", 10, "bold"), command=self.aktifkan_scan_kamera, bg="#d4d0c8", bd=2, relief="raised").pack(side="left", padx=15, pady=10)
        
        self.KOTAK_PESAN = tk.Frame(self.tab_absensi, bg="#b5b2ad", bd=3, relief="groove")
        self.KOTAK_PESAN.pack(fill="both", expand=True, padx=15, pady=(0, 20))
        
        self.lbl_respons = tk.Label(self.KOTAK_PESAN, text="Gesekan kartu absensi Anda\nmelalui mesin barcode", font=("Arial", 18, "bold"), fg="black", bg="#b5b2ad", justify="center")
        self.lbl_respons.pack(fill="x", pady=(40, 10))
        
        tk.Frame(self.KOTAK_PESAN, bg="#7a7772", height=2).pack(fill="x", padx=40, pady=10)
        
        self.lbl_tanggal = tk.Label(self.KOTAK_PESAN, text="", font=("Arial", 16, "bold"), fg="black", bg="#b5b2ad")
        self.lbl_tanggal.pack(pady=2)
        self.lbl_jam = tk.Label(self.KOTAK_PESAN, text="", font=("Arial", 20, "bold"), fg="black", bg="#b5b2ad")
        self.lbl_jam.pack(pady=(0, 20))
        
        self.perbarui_jam()

    def perbarui_jam(self):
        waktu_sekarang = datetime.now()
        self.lbl_tanggal.config(text=waktu_sekarang.strftime("%A, %d %B %Y"))
        self.lbl_jam.config(text=waktu_sekarang.strftime("%H:%M:%S WIB"))
        self.root.after(1000, self.perbarui_jam)

    def reset_tampilan(self):
        self.lbl_respons.config(text="Gesekan Kartu Absensi Anda\nMelalui Mesin Barcode", fg="black")

    def aktifkan_scan_kamera(self):
        if self.is_camera_on: return
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showwarning("Kamera Error", "Perangkat keras kamera tidak ditemukan!")
            return
        self.is_camera_on = True
        self.lbl_respons.config(text="Dekatkan barcode ke arah lensa kamera.", fg="blue")
        self.loop_proses_kamera()

    def loop_proses_kamera(self):
        if not self.is_camera_on: return
        ret, frame = self.cap.read()
        if ret:
            barcodes = decode(frame)
            for barcode_data_raw in barcodes:
                data_barcode = barcode_data_raw.data.decode('utf-8').strip()
                if data_barcode:
                    self.input_barcode.delete(0, tk.END)
                    self.input_barcode.insert(0, data_barcode)
                    
                    self.is_camera_on = False
                    self.cap.release()
                    cv2.destroyAllWindows()
                    
                    self.proses_absensi(None)
                    return
            cv2.imshow('Pemindai Barcode Kamera', frame)
            
        if cv2.waitKey(1) & 0xFF == ord('q') or cv2.getWindowProperty('Pemindai Barcode Kamera', cv2.WND_PROP_VISIBLE) < 1 if 'Pemindai Barcode Kamera' in locals() else False:
            self.is_camera_on = False
            self.cap.release()
            cv2.destroyAllWindows()
            self.reset_tampilan()
            return
            
        if self.is_camera_on:
            self.root.after(10, self.loop_proses_kamera)

    def proses_absensi(self, event):
        """Logika Utama Absensi disertai Aturan Validasi Jeda Waktu (Lockout Duration)."""
        if self.after_id is not None:
            self.root.after_cancel(self.after_id)
            self.after_id = None

        master_karyawan = self.ambil_master_karyawan()
        id_scan = self.input_barcode.get().strip()
        self.input_barcode.delete(0, tk.END) 
        
        if id_scan not in master_karyawan:
            self.lbl_respons.config(text="ID CARD TIDAK DIKENALI!\nKaryawan Belum Terdaftar.", fg="red")
            self.after_id = self.root.after(5000, self.reset_tampilan)
            return
            
        nama = master_karyawan[id_scan]
        waktu_scan = datetime.now()
        tanggal_hari_ini = waktu_scan.strftime("%Y-%m-%d")
        waktu_str = waktu_scan.strftime("%H:%M:%S")
        
        # Ambil data lengkap riwayat log hari ini
        log_hari_ini = self.cari_atau_buat_log_hari_ini(id_scan, nama, tanggal_hari_ini)
        row_index = log_hari_ini["row_index"]
        status_hari_ini = log_hari_ini["status"]
        
        # Helper internal untuk mengubah teks jam dari Excel kembali menjadi objek datetime objek hari ini
        def dapatkan_datetime_log(jam_str):
            if jam_str and jam_str != "-":
                return datetime.strptime(f"{tanggal_hari_ini} {jam_str}", "%Y-%m-%d %H:%M:%S")
            return None

        # ---------------------------------------------------------------------
        # KEPUTUSAN LOGIKA & PEMBATASAN WAKTU (STATE MACHINE WITH LOCKOUT)
        # ---------------------------------------------------------------------
        if status_hari_ini == "BELUM_MASUK":
            if self.perbarui_log_absensi(row_index, kolom_ke=5, status_baru="SUDAH_MASUK", waktu_scan=waktu_scan):
                self.lbl_respons.config(text=f"Selamat Datang, {nama}!\n\nJam Masuk Dicatat:\n {waktu_str}", fg="darkgreen")
            
        elif status_hari_ini == "SUDAH_MASUK":
            waktu_masuk_dt = dapatkan_datetime_log(log_hari_ini["waktu_masuk"])
            
            # Validasi jeda 1 jam dari Masuk ke Istirahat
            if waktu_masuk_dt and waktu_scan < waktu_masuk_dt + timedelta(seconds=10): # Atur batas jeda waktu masuk ke istirahat harusnya 1 jam
                sisa_waktu = (waktu_masuk_dt + timedelta(seconds=10)) - waktu_scan  # Atur batas jeda waktu masuk ke istirahat harusnya 1 jam
                menit_sisa = int(sisa_waktu.total_seconds() / 60)
                self.lbl_respons.config(text=f"\n\n{nama} Anda baru Saja Absen Masuk Pada Pukul\n{waktu_masuk_dt}", fg="red")
            else:
                if self.perbarui_log_absensi(row_index, kolom_ke=6, status_baru="ISTIRAHAT", waktu_scan=waktu_scan):
                    self.lbl_respons.config(text=f"☕ Selamat Istirahat, {nama}.\n\nJam Keluar Istirahat: {waktu_str}", fg="darkgreen")
            
        elif status_hari_ini == "ISTIRAHAT":
            waktu_istirahat_dt = dapatkan_datetime_log(log_hari_ini["waktu_istirahat"])
            
            # Validasi jeda 5 menit dari Istirahat ke Kembali Kerja
            if waktu_istirahat_dt and waktu_scan < waktu_istirahat_dt + timedelta(seconds=30): # Atur batas jeda waktu istirahat harunya 5 menit
                sisa_waktu = (waktu_istirahat_dt + timedelta(seconds=30)) - waktu_scan # Atur batas jeda waktu istirahat harunya 5 menit
                detik_sisa = int(sisa_waktu.total_seconds()/60)
                self.lbl_respons.config(text=f"\n{nama} Anda Sudah Istirahat pada Pukul\n{waktu_istirahat_dt}\n Anda Bisa scan Masuk\n{detik_sisa} menit lagi", fg="red")
            else:
                if self.perbarui_log_absensi(row_index, kolom_ke=7, status_baru="KEMBALI_KERJA", waktu_scan=waktu_scan):
                    self.lbl_respons.config(text=f"💼 Selamat Bekerja Kembali, {nama}!\n\nJam Kembali Kerja: {waktu_str}", fg="darkgreen")
            
        elif status_hari_ini == "KEMBALI_KERJA":
            waktu_istirahat_dt = dapatkan_datetime_log(log_hari_ini["waktu_istirahat"])
            
            # Validasi jeda 1 jam sejak *mulai istirahat* sebelum diperbolehkan scan pulang
            if waktu_istirahat_dt and waktu_scan < waktu_istirahat_dt + timedelta(seconds=30): # Atur batas waktu istirahat harunya 1 jam
                sisa_waktu = (waktu_istirahat_dt + timedelta(seconds=30)) - waktu_scan   # Atur batas waktu istirahat harunya 1 jam
                menit_sisa = int(sisa_waktu.total_seconds() / 60)
                self.lbl_respons.config(text=f"\n{nama} Anda baru Saja Kembali Istirahat Pada Pukul\n{waktu_istirahat_dt}", fg="red")
            else:
                if self.perbarui_log_absensi(row_index, kolom_ke=8, status_baru="PULANG", waktu_scan=waktu_scan):
                    self.lbl_respons.config(text=f"Terima Kasih Atas Kerja Kerasnya, {nama}!\n\nJam Pulang Dicatat:\n⏰ {waktu_str}", fg="blue")
            
        elif status_hari_ini == "PULANG":
            self.lbl_respons.config(text=f"ℹ️ Halo {nama},\nAnda sudah menyelesaikan absen hari ini.", fg="black")

        self.after_id = self.root.after(5000, self.reset_tampilan)

    # =========================================================================
    # MENU 2 & 3 (TETAP SAMA SEPERTI VERSI MULTI-SHEET SEBELUMNYA)
    # =========================================================================
    def buat_gui_tab_generator(self):
        frame_form = tk.Frame(self.tab_generator, bg="#d4d0c8", bd=2, relief="groove")
        frame_form.pack(fill="both", expand=True, padx=20, pady=20)
        tk.Label(frame_form, text="Nomor ID Karyawan Baru (Permanen):", font=("Arial", 10, "bold"), bg="#d4d0c8").pack(anchor="w", padx=30, pady=(30, 2))
        self.entry_gen_id = tk.Entry(frame_form, font=("Arial", 11), width=40, bd=2, relief="sunken")
        self.entry_gen_id.pack(padx=30, pady=(0, 15))
        tk.Label(frame_form, text="Nama Lengkap Karyawan Baru:", font=("Arial", 10, "bold"), bg="#d4d0c8").pack(anchor="w", padx=30, pady=(0, 2))
        self.entry_gen_nama = tk.Entry(frame_form, font=("Arial", 11), width=40, bd=2, relief="sunken")
        self.entry_gen_nama.pack(padx=30, pady=(0, 30))
        tk.Button(frame_form, text="⚙️ Daftarkan Karyawan Tetap & Bikin Barcode", font=("Arial", 11, "bold"), command=self.proses_generate_barcode, bg="#d4d0c8", bd=2, relief="raised", padx=15, pady=5).pack()
        self.lbl_status_gen = tk.Label(frame_form, text="Status: Siap mendaftarkan karyawan jangka panjang.", font=("Arial", 9, "italic"), bg="#d4d0c8", fg="gray")
        self.lbl_status_gen.pack(side="bottom", pady=10)

    def proses_generate_barcode(self):
        id_baru = self.entry_gen_id.get().strip()
        nama_baru = self.entry_gen_nama.get().strip()
        if not id_baru or not nama_baru:
            messagebox.showwarning("Data Kosong", "Seluruh kolom data wajib diisi!")
            return
        master_karyawan = self.ambil_master_karyawan()
        if id_baru in master_karyawan:
            messagebox.showwarning("Duplikasi", "Nomor ID Karyawan tersebut sudah bekerja di perusahaan ini!")
            return
        try:
            barcode_class = barcode.get_barcode_class('code128')
            barcode_obj = barcode_class(id_baru, writer=ImageWriter())
            nama_file_aman = "".join([c for c in nama_baru if c.isalpha() or c.isdigit() or c==' ']).rstrip()
            path_file = os.path.join(self.folder_barcode, f"{id_baru}_{nama_file_aman}")
            barcode_obj.save(path_file, options={'write_text': True, 'background': 'white', 'foreground': 'black', 'text_distance': 5.0, 'font_size': 11})
            wb = openpyxl.load_workbook(self.nama_file_excel)
            ws_karyawan = wb["Master_Karyawan"]
            ws_karyawan.append([id_baru, nama_baru, datetime.now().strftime("%Y-%m-%d")])
            wb.save(self.nama_file_excel)
            self.entry_gen_id.delete(0, tk.END)
            self.entry_gen_nama.delete(0, tk.END)
            self.lbl_status_gen.config(text=f"✅ Sukses mendaftarkan {nama_baru} ke database perusahaan!", fg="green")
            messagebox.showinfo("Sukses", f"Karyawan baru terdaftar permanen!\nBarcode disimpan di folder '{self.folder_barcode}'")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal memproses data: {str(e)}")

    def buat_gui_tab_rekap(self):
        frame_rekap = tk.Frame(self.tab_rekap, bg="#d4d0c8", bd=2, relief="groove")
        frame_rekap.pack(fill="both", expand=True, padx=20, pady=20)
        tk.Label(frame_rekap, text="📊 Panel Ekstraksi Log Riwayat Kehadiran", font=("Arial", 12, "bold"), bg="#d4d0c8").pack(pady=(20, 5))
        tk.Label(frame_rekap, text="Pilih bulan untuk menarik seluruh rekapan log multi-hari:", font=("Arial", 10), bg="#d4d0c8").pack(pady=(0, 20))
        frame_pilih = tk.Frame(frame_rekap, bg="#d4d0c8")
        frame_pilih.pack(pady=10)
        tk.Label(frame_pilih, text="Bulan Laporan :", font=("Arial", 10, "bold"), bg="#d4d0c8").pack(side="left", padx=5)
        self.combo_bulan = ttk.Combobox(frame_pilih, values=["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"], state="readonly", width=15)
        self.combo_bulan.pack(side="left", padx=5)
        bulan_sekarang = datetime.now().month - 1
        self.combo_bulan.current(bulan_sekarang)
        tk.Button(frame_rekap, text="📥 Download Laporan Historis Bulanan", font=("Arial", 11, "bold"), command=self.proses_cetak_rekap_bulanan, bg="#d4d0c8", bd=2, relief="raised", padx=15, pady=8).pack(pady=30)
        self.lbl_status_rekap = tk.Label(frame_rekap, text="Sistem akan mem-filter log database berdasarkan bulan yang dipilih.", font=("Arial", 9, "italic"), bg="#d4d0c8", fg="gray")
        self.lbl_status_rekap.pack(side="bottom", pady=15)

    def proses_cetak_rekap_bulanan(self):
        bulan_dipilih = self.combo_bulan.get()
        mapping_bulan = {"Januari": "01", "Februari": "02", "Maret": "03", "April": "04", "Mei": "05", "Juni": "06", "Juli": "07", "Agustus": "08", "September": "09", "Oktober": "10", "November": "11", "Desember": "12"}
        kode_bulan_target = mapping_bulan[bulan_dipilih]
        wb_rekap = openpyxl.Workbook()
        ws_rekap = wb_rekap.active
        ws_rekap.title = f"Log_{bulan_dipilih}"
        ws_rekap.merge_cells('A1:H1')
        ws_rekap['A1'] = f"REKAPITULASI RIWAYAT KEHADIRAN KARYAWAN - BULAN {bulan_dipilih.upper()} {datetime.now().year}"
        ws_rekap['A1'].font = Font(name='Arial', size=14, bold=True, color='0A246A')
        ws_rekap['A1'].alignment = Alignment(horizontal='center')
        ws_rekap.append([]) 
        headers = ["Tanggal", "ID Karyawan", "Nama Karyawan", "Status Terakhir", "Jam Masuk", "Jam Istirahat", "Kembali Kerja", "Jam Pulang"]
        ws_rekap.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws_rekap.cell(row=3, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0A246A', end_color='0A246A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        wb_utama = openpyxl.load_workbook(self.nama_file_excel, data_only=True)
        ws_log = wb_utama["Log_Absensi"]
        ada_data = False
        for row in ws_log.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                tanggal_str = str(row[0]).strip()
                try:
                    bulan_log = tanggal_str.split("-")[1]
                    if bulan_log == kode_bulan_target:
                        ws_rekap.append([row[0], row[1], row[2], row[3], row[4] if row[4] else "-", row[5] if row[5] else "-", row[6] if row[6] else "-", row[7] if row[7] else "-"])
                        ada_data = True
                except IndexError:
                    continue
        if not ada_data:
            messagebox.showinfo("Informasi", f"Tidak ditemukan riwayat log absensi pada bulan {bulan_dipilih}.")
            return
        for col in ws_rekap.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_rekap.column_dimensions[col_letter].width = max(max_len + 3, 15)
        nama_file_rekap = f"Laporan_Absensi_Historis_{bulan_dipilih}_{datetime.now().year}.xlsx"
        try:
            wb_rekap.save(nama_file_rekap)
            messagebox.showinfo("Rekap Berhasil", f"Laporan bulanan sukses diekspor!\n\nFile: {nama_file_rekap}")
        except PermissionError:
            messagebox.showerror("Gagal Menyimpan", f"File berkas '{nama_file_rekap}' sedang dibuka di Microsoft Excel! Silakan tutup terlebih dahulu.")

if __name__ == "__main__":
    root = tk.Tk()
    app = SistemHRDAndAbsensi(root)
    root.mainloop()